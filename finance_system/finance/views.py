import re
from django.db.models import Q, Sum
from django.db.models.functions import TruncMonth
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.utils import timezone
from datetime import date, timedelta
import calendar
import json

from django.http import JsonResponse, HttpResponse
from .forms import CustomUserCreationForm, CustomAuthenticationForm, FinancialGoalForm, CategoryForm, ReceiptUploadForm, ProfileUpdateForm
from .models import Category, Transaction, FinancialGoal, GoalContribution, Account, Family, FamilyMember, Notification, FamilyInvitation, CustomUser
from .utils.receipt_ocr import extract_receipt_data


def index(request):
    """Главная страница - редирект для авторизованных"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'finance/index.html')


def features(request):
    """Страница возможностей"""
    return render(request, 'finance/features.html')


def pricing(request):
    """Страница тарифов"""
    return render(request, 'finance/pricing.html')


def contact(request):
    """Страница контактов"""
    return render(request, 'finance/contact.html')


def auth_view(request):
    """Единая страница аутентификации"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    return render(request, 'finance/auth.html', {
        'login_form': CustomAuthenticationForm(),
        'register_form': CustomUserCreationForm(),
        'active_tab': 'login',
    })


def handle_auth(request):
    """Обработка всех форм аутентификации"""
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'login':
            login_form = CustomAuthenticationForm(request, data=request.POST)
            if login_form.is_valid():
                username = login_form.cleaned_data.get('username')
                password = login_form.cleaned_data.get('password')
                user = authenticate(username=username, password=password)

                if user is not None:
                    login(request, user)
                    messages.success(request, f'Добро пожаловать, {user.username}!')
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Неверное имя пользователя или пароль.')
                    return render(request, 'finance/auth.html', {
                        'login_form': login_form,
                        'register_form': CustomUserCreationForm(),
                        'active_tab': 'login',
                    })
            else:
                return render(request, 'finance/auth.html', {
                    'login_form': login_form,
                    'register_form': CustomUserCreationForm(),
                    'active_tab': 'login',
                })

        elif form_type == 'register':
            register_form = CustomUserCreationForm(request.POST)
            if register_form.is_valid():
                user = register_form.save()
                login(request, user)
                messages.success(request, 'Регистрация успешна! Добро пожаловать!')
                return redirect('dashboard')
            else:
                return render(request, 'finance/auth.html', {
                    'login_form': CustomAuthenticationForm(),
                    'register_form': register_form,
                    'active_tab': 'register',
                })

    return redirect('auth')


def login_view(request):
    """Отдельный обработчик для входа"""
    return handle_auth(request)


def register_view(request):
    """Отдельный обработчик для регистрации"""
    return handle_auth(request)


@login_required
def logout_view(request):
    """Выход из системы"""
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы.')
    return redirect('index')


def redirect_to_admin(request):
    """Редирект /site-admin/ на кастомную админку /admin/ (только для is_staff)."""
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_dashboard')
    messages.error(request, 'Доступ только для администратора.')
    return redirect('dashboard')


@login_required
def site_admin_panel(request):
    """Кастомная админ-панель сайта — только для пользователей с ролью admin (is_staff)."""
    if not request.user.is_staff:
        messages.error(request, 'Доступ только для администратора.')
        return redirect('dashboard')
    from django.utils import formats
    users_count = CustomUser.objects.count()
    families_count = Family.objects.count()
    goals_count = FinancialGoal.objects.count()
    goals_active = FinancialGoal.objects.filter(status='active').count()
    # График пополнений по месяцам (все цели)
    monthly = GoalContribution.objects.all().annotate(
        month=TruncMonth('contributed_at')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    chart_labels = []
    chart_data = []
    for row in monthly:
        chart_labels.append(formats.date_format(row['month'], 'Y-m') if row['month'] else '')
        chart_data.append(float(row['total']))
    return render(request, 'finance/site_admin_panel.html', {
        'users_count': users_count,
        'families_count': families_count,
        'goals_count': goals_count,
        'goals_active': goals_active,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_data_json': json.dumps(chart_data),
    })


@login_required
def dashboard(request):
    """Панель управления со всеми вкладками"""
    user_goals = FinancialGoal.objects.filter(user=request.user)
    user_families_ids = Family.objects.filter(
        Q(created_by=request.user) | Q(members__user=request.user)
    ).values_list('id', flat=True)
    family_goals = FinancialGoal.objects.filter(family_id__in=user_families_ids)
    transactions = Transaction.objects.filter(user=request.user).order_by('-date', '-created_at')
    categories = Category.objects.filter(
        Q(owner=request.user) | Q(is_system=True),
        type='expense'
    ).distinct()
    # Порядок для чека: Еда, Здоровье, Кафе и рестораны, затем остальные
    from django.db.models import Case, When, Value, IntegerField
    PREFERRED_NAMES = ['Еда', 'Здоровье', 'Кафе и рестораны']
    order_cases = [When(name=name, then=Value(i)) for i, name in enumerate(PREFERRED_NAMES)]
    categories = categories.annotate(
        _sort=Case(*order_cases, default=Value(100), output_field=IntegerField())
    ).order_by('_sort', 'name')

    # Статистика для вкладки Обзор
    total_goals = user_goals.count() + family_goals.count()
    active_goals = user_goals.filter(status='active').count() + family_goals.filter(status='active').count()

    total_current_amount = sum([float(g.current_amount) for g in user_goals]) + sum([float(g.current_amount) for g in family_goals])
    total_target_amount = sum([float(g.target_amount) for g in user_goals]) + sum([float(g.target_amount) for g in family_goals])

    # Статистика расходов
    category_stats = {}
    total_expenses = 0
    transaction_count = transactions.count()

    for transaction in transactions.filter(type='expense'):
        if transaction.category:
            cat_name = transaction.category.name
            if cat_name not in category_stats:
                category_stats[cat_name] = {
                    'amount': 0,
                    'count': 0,
                    'color': transaction.category.color
                }
            category_stats[cat_name]['amount'] += float(transaction.amount)
            category_stats[cat_name]['count'] += 1
            total_expenses += float(transaction.amount)

    # Рассчитываем проценты
    for cat in category_stats.values():
        if total_expenses > 0:
            cat['percentage'] = (cat['amount'] / total_expenses) * 100
        else:
            cat['percentage'] = 0

    # График расходов по месяцам (последние 12 месяцев)
    from django.db.models.functions import TruncMonth
    expense_monthly = Transaction.objects.filter(
        user=request.user, type='expense'
    ).annotate(month=TruncMonth('date')).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    expense_chart_months = []
    expense_chart_data = []
    today = timezone.now().date()
    for i in range(11, -1, -1):
        y, m = today.year, today.month - i
        while m <= 0:
            m += 12
            y -= 1
        month_str = date(y, m, 1).strftime('%Y-%m')
        expense_chart_months.append(month_str)
        val = 0
        for row in expense_monthly:
            if row.get('month'):
                row_str = row['month'].strftime('%Y-%m') if hasattr(row['month'], 'strftime') else str(row['month'])[:7]
                if row_str == month_str:
                    val = float(row.get('total') or 0)
                    break
        expense_chart_data.append(val)

    user_families = Family.objects.filter(
        Q(created_by=request.user) | Q(members__user=request.user)
    ).distinct()

    user_accounts = Account.objects.filter(
        Q(owner=request.user) | Q(family__members__user=request.user),
        is_active=True
    ).distinct()

    context = {
        'goals': user_goals,
        'transactions': transactions[:50],
        'categories': categories,
        'user_families': user_families,
        'user_accounts': user_accounts,
        'total_goals': total_goals,
        'active_goals': active_goals,
        'total_current_amount': total_current_amount,
        'total_target_amount': total_target_amount,
        'total_expenses': total_expenses,
        'transaction_count': transaction_count,
        'category_count': categories.count(),
        'system_category_count': categories.filter(is_system=True).count(),
        'category_stats': category_stats,
        'expense_chart_months': json.dumps(expense_chart_months),
        'expense_chart_data': json.dumps(expense_chart_data),
        'expense_progress_pct': min(100, int(total_expenses) / max(100000, int(total_expenses)) * 100) if total_expenses else 0,
        'active_tab': request.GET.get('tab', 'overview'),
    }

    return render(request, 'finance/dashboard.html', context)


@login_required
def categories_redirect(request):
    """Страница категорий - редирект на dashboard"""
    return redirect(reverse('dashboard') + '?tab=categories')


@login_required
def transactions_redirect(request):
    """Страница транзакций - редирект на dashboard"""
    return redirect(reverse('dashboard') + '?tab=transactions')


@login_required
def goals_redirect(request):
    """Страница целей - редирект на dashboard"""
    return redirect(reverse('dashboard') + '?tab=goals')


@login_required
def create_category(request):
    """Создание новой категории"""
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            # Создаем категорию вручную
            import random
            colors = [
                '#FF6B6B', '#4ECDC4', '#FFD166', '#06D6A0', '#118AB2',
                '#EF476F', '#1B9AAA', '#06BCC1', '#F86624', '#662E9B',
                '#2A9D8F', '#E9C46A', '#F4A261', '#E76F51'
            ]

            category = Category.objects.create(
                name=name,
                owner=request.user,
                type='expense',
                color=random.choice(colors)
            )
            messages.success(request, 'Категория создана!')
        else:
            messages.error(request, 'Введите название категории')
        return redirect(reverse('dashboard') + '?tab=categories')
    return redirect(reverse('dashboard') + '?tab=categories')


@login_required
def delete_category(request, category_id):
    """Удаление категории (только своей, не системной)."""
    category = get_object_or_404(Category, id=category_id)
    if category.is_system or (category.owner_id and category.owner_id != request.user.id):
        messages.error(request, 'Нельзя удалить эту категорию')
        return redirect(reverse('dashboard') + '?tab=categories')
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Категория удалена')
        return redirect(reverse('dashboard') + '?tab=categories')
    return render(request, 'finance/category_delete.html', {'category': category})


@login_required
def scan_receipt(request):
    """Сканирование чека (QR + OCR): возвращает JSON с суммой, магазином и категорией."""
    if request.method != 'POST' or not request.FILES.get('receipt_image'):
        return JsonResponse({'ok': False, 'error': 'Загрузите изображение чека'})
    data = {'amount': None, 'merchant': '', 'suggested_category': None, 'date': None}
    try:
        image = request.FILES['receipt_image']
        data = extract_receipt_data(image)
    except Exception as e:
        pass
    category_id = None
    if data.get('suggested_category'):
        cat = Category.objects.filter(
            Q(owner=request.user) | Q(is_system=True),
            type='expense',
            name=data['suggested_category']
        ).first()
        if cat:
            category_id = str(cat.id)
    if not category_id:
        cat = Category.objects.filter(
            Q(owner=request.user) | Q(is_system=True),
            type='expense'
        ).first()
        if cat:
            category_id = str(cat.id)
    resp = {
        'ok': True,
        'amount': data.get('amount'),
        'merchant': data.get('merchant') or '',
        'suggested_category': data.get('suggested_category') or '',
        'category_id': category_id,
    }
    if data.get('date'):
        resp['date'] = data['date']
    return JsonResponse(resp)


@login_required
def upload_receipt(request):
    """Загрузка чека"""
    if request.method == 'POST':
        post_data = request.POST.copy()
        if not post_data.get('category'):
            default_cat = Category.objects.filter(
                Q(owner=request.user) | Q(is_system=True),
                type='expense'
            ).first()
            if default_cat:
                post_data['category'] = str(default_cat.id)
        if not post_data.get('date'):
            from datetime import datetime
            post_data['date'] = timezone.now().strftime('%Y-%m-%dT%H:%M:%S')
        form = ReceiptUploadForm(post_data, request.FILES, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.type = 'expense'
            transaction.currency = 'RUB'
            transaction.created_via = 'scan'
            transaction.merchant = (request.POST.get('description') or transaction.merchant or '')[:200]

            default_account = Account.objects.filter(owner=request.user, is_active=True).first()
            if not default_account:
                default_account = Account.objects.create(
                    owner=request.user, name='Основной счёт', account_type='debit',
                    ownership='personal', currency='RUB', is_active=True
                )
            transaction.account = default_account

            transaction.save()

            messages.success(request, 'Чек успешно загружен!')
            return redirect(reverse('dashboard') + '?tab=upload')
        else:
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f'{field}: {e}')
    return redirect(reverse('dashboard') + '?tab=upload')


def _get_or_create_default_account(user):
    """Возвращает счёт пользователя или создаёт основной счёт."""
    account = Account.objects.filter(owner=user, is_active=True).first()
    if not account:
        account = Account.objects.create(
            owner=user, name='Основной счёт', account_type='debit',
            ownership='personal', currency='RUB', is_active=True
        )
    return account


@login_required
def add_transaction(request):
    """Ручное добавление транзакции"""
    if request.method != 'POST':
        return redirect(reverse('dashboard') + '?tab=transactions')
    amount = request.POST.get('amount')
    trans_type = 'expense'
    cat_id = request.POST.get('category')
    account_id = request.POST.get('account')
    date_str = request.POST.get('date')
    merchant = (request.POST.get('merchant') or '')[:200]
    try:
        amount_val = float(str(amount).replace(',', '.').replace(' ', ''))
        if amount_val <= 0:
            raise ValueError('Сумма должна быть больше 0')
    except (ValueError, TypeError):
        messages.error(request, 'Введите корректную сумму')
        return redirect(reverse('dashboard') + '?tab=transactions')
    account = None
    if account_id:
        account = Account.objects.filter(
            Q(owner=request.user) | Q(family__members__user=request.user),
            id=account_id, is_active=True
        ).first()
    if not account:
        account = _get_or_create_default_account(request.user)
    category = None
    if cat_id:
        category = Category.objects.filter(
            Q(owner=request.user) | Q(is_system=True),
            id=cat_id
        ).first()
    trans_date = timezone.now()
    if date_str:
        from datetime import datetime
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                trans_date = timezone.make_aware(datetime.strptime(date_str.strip(), fmt))
                break
            except (ValueError, TypeError):
                continue
    Transaction.objects.create(
        user=request.user, account=account, category=category,
        amount=amount_val, type=trans_type, currency='RUB',
        description='', merchant=merchant or None,
        date=trans_date, created_via='manual'
    )
    messages.success(request, 'Транзакция добавлена')
    return redirect(reverse('dashboard') + '?tab=transactions')


@login_required
def import_transactions_excel(request):
    """Импорт транзакций из Excel файла"""
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        messages.error(request, 'Выберите Excel файл для загрузки')
        return redirect(reverse('dashboard') + '?tab=transactions')
    try:
        import openpyxl
        from datetime import datetime
    except ImportError:
        messages.error(request, 'Установите openpyxl: pip install openpyxl')
        return redirect(reverse('dashboard') + '?tab=transactions')
    account_id = request.POST.get('account')
    account = None
    if account_id:
        account = Account.objects.filter(
            Q(owner=request.user) | Q(family__members__user=request.user),
            id=account_id, is_active=True
        ).first()
    if not account:
        account = _get_or_create_default_account(request.user)
    categories_qs = Category.objects.filter(
        Q(owner=request.user) | Q(is_system=True), type='expense'
    )
    categories_by_name = {c.name.lower().strip(): c for c in categories_qs}
    created = 0
    errors = []
    wb = openpyxl.load_workbook(request.FILES['excel_file'], read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        try:
            date_cell = row[0]
            amount_cell = row[1]
            cat_cell = (row[2] if len(row) > 2 else '') or ''
            merchant_cell = (row[3] if len(row) > 3 else '') or ''
            if date_cell is None or amount_cell is None:
                continue
            if hasattr(date_cell, 'date'):
                trans_date = timezone.make_aware(
                    datetime.combine(date_cell.date() if hasattr(date_cell, 'date') else date_cell, datetime.min.time())
                )
            else:
                date_str = str(date_cell).strip()
                trans_date = timezone.now()
                for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        trans_date = timezone.make_aware(datetime.strptime(date_str[:10], fmt))
                        break
                    except (ValueError, TypeError):
                        continue
            amount_val = float(str(amount_cell).replace(',', '.').replace(' ', ''))
            if amount_val <= 0:
                continue
            trans_type = 'expense'
            category = None
            if cat_cell:
                cat_name = str(cat_cell).strip().lower()
                category = categories_by_name.get(cat_name)
                if not category:
                    for c in categories_qs:
                        if c.name.lower().strip() == cat_name:
                            category = c
                            break
            Transaction.objects.create(
                user=request.user, account=account, category=category,
                amount=amount_val, type=trans_type, currency='RUB',
                description='',
                merchant=str(merchant_cell).strip()[:200] or None,
                date=trans_date, created_via='import'
            )
            created += 1
        except Exception as e:
            errors.append(f'Строка {i}: {str(e)[:80]}')
    wb.close()
    if created > 0:
        messages.success(request, f'Импортировано {created} транзакций')
    if errors:
        messages.warning(request, f'Ошибки: {"; ".join(errors[:3])}{"..." if len(errors) > 3 else ""}')
    return redirect(reverse('dashboard') + '?tab=transactions')


@login_required
def download_transactions_example(request):
    """Скачивание примера Excel для импорта транзакций"""
    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        messages.error(request, 'openpyxl не установлен')
        return redirect(reverse('dashboard') + '?tab=transactions')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Транзакции'
    headers = ['Дата', 'Сумма', 'Категория', 'Магазин/Продавец']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    example_rows = [
        ['2025-01-15', 1500, 'Еда', 'Пятёрочка'],
        ['2025-01-16', 350, 'Кафе и рестораны', 'Шоколадница'],
        ['2025-01-17', 2500, '', ''],
    ]
    for row_idx, row_data in enumerate(example_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="transactions_example.xlsx"'
    return response


@login_required
def create_goal(request):
    """Создание цели: личной или семейной (выбор семьи)."""
    if request.method == 'POST':
        form = FinancialGoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal_scope = request.POST.get('goal_scope', 'personal')
            family_id = request.POST.get('family_id') or None

            if goal_scope == 'family' and family_id:
                family = get_object_or_404(Family, id=family_id)
                if family.created_by != request.user and not family.members.filter(user=request.user).exists():
                    messages.error(request, 'Нет доступа к этой семье')
                    return redirect('dashboard')
                is_creator_or_admin = family.created_by == request.user or family.members.filter(user=request.user, role='admin').exists()
                can_create = is_creator_or_admin or getattr(family, 'members_can_create_goals', True)
                if not can_create:
                    messages.error(request, 'В этой семье только создатель/админ могут создавать цели. Включите опцию в настройках семьи.')
                    return redirect('family_detail', family_id=family_id)
                goal.family = family
                goal.user = None
            else:
                goal.user = request.user
                goal.family = None

            goal.replenishment_frequency = request.POST.get('replenishment_frequency', '').strip() or ''
            goal.current_amount = 0
            goal.status = 'active'
            goal.start_date = date.today()
            if goal.target_amount and goal.target_amount > 0:
                goal.progress_percentage = 0
            else:
                goal.progress_percentage = 0
            goal.save()
            messages.success(request, 'Цель успешно создана!')
            if goal.family_id:
                return redirect('family_detail', family_id=goal.family_id)
            return redirect(reverse('dashboard') + '?tab=goals')
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")
    return redirect(reverse('dashboard') + '?tab=goals')


@login_required
def edit_goal(request, goal_id):
    """Редактирование цели (личной или семейной)."""
    goal = get_object_or_404(FinancialGoal, id=goal_id)
    if not _user_can_edit_goal(request, goal):
        messages.error(request, 'Нет доступа к этой цели')
        return redirect('dashboard')

    redirect_after = reverse('dashboard') + '?tab=goals'
    if goal.family_id:
        redirect_after = reverse('family_detail', kwargs={'family_id': goal.family_id})

    if request.method == 'POST':
        form = FinancialGoalForm(request.POST, instance=goal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Цель обновлена!')
            return redirect(redirect_after)
    else:
        form = FinancialGoalForm(instance=goal)
    return render(request, 'finance/edit_goal.html', {'form': form, 'goal': goal, 'redirect_after': redirect_after})


@login_required
def delete_goal(request, goal_id):
    """Удаление цели (личной или семейной)."""
    goal = get_object_or_404(FinancialGoal, id=goal_id)
    if not _user_can_edit_goal(request, goal):
        messages.error(request, 'Нет доступа к этой цели')
        return redirect('dashboard')

    redirect_after = reverse('dashboard') + '?tab=goals'
    if goal.family_id:
        redirect_after = reverse('family_detail', kwargs={'family_id': goal.family_id})

    if request.method == 'POST':
        goal.delete()
        messages.success(request, 'Цель удалена!')
        return redirect(redirect_after)
    return render(request, 'finance/delete_goal.html', {'goal': goal, 'redirect_after': redirect_after})


def _user_can_edit_goal(request, goal):
    """Проверка доступа: личная цель пользователя или член семьи для семейной цели."""
    if goal.user_id and goal.user_id == request.user.id:
        return True
    if goal.family_id:
        return goal.family.created_by_id == request.user.id or goal.family.members.filter(user=request.user).exists()
    return False


@login_required
def add_money_to_goal(request, goal_id):
    """Добавление или вычитание денег из цели (личной или семейной)."""
    goal = get_object_or_404(FinancialGoal, id=goal_id)
    if not _user_can_edit_goal(request, goal):
        messages.error(request, 'Нет доступа к этой цели')
        return redirect('dashboard')

    redirect_to = reverse('dashboard') + '?tab=goals'
    if goal.family_id:
        redirect_to = reverse('family_detail', kwargs={'family_id': goal.family_id})

    if request.method == 'POST':
        try:
            amount_str = (request.POST.get('amount') or '').strip()
            amount_str = amount_str.replace(' ', '').replace('\xa0', '').replace('\u202f', '').replace('\u2009', '')
            if '.' in amount_str or ',' in amount_str:
                messages.error(request, 'Введите сумму целым числом без копеек (например: 1000)')
                return redirect(redirect_to)
            amount_str = re.sub(r'[^\d]', '', amount_str)
            if not amount_str:
                messages.error(request, 'Введите сумму')
                return redirect(redirect_to)

            amount = int(amount_str)
            if amount <= 0:
                messages.error(request, 'Сумма должна быть больше 0')
                return redirect(redirect_to)

            action = (request.POST.get('action') or 'add').strip().lower()

            if action == 'subtract':
                goal.current_amount = max(0, goal.current_amount - amount)
                if goal.status == 'completed' and goal.current_amount < goal.target_amount:
                    goal.status = 'active'
                messages.success(request, f'Списано {amount} ₽ с цели «{goal.name}»')
            else:
                goal.current_amount += amount
                GoalContribution.objects.create(
                    goal=goal,
                    amount=amount,
                    user=request.user,
                    contributed_at=timezone.now(),
                )
                goal.last_replenishment_at = date.today()
                if goal.current_amount >= goal.target_amount:
                    goal.current_amount = goal.target_amount
                    goal.status = 'completed'
                    messages.success(request, f'Поздравляем! Цель «{goal.name}» выполнена!')
                else:
                    messages.success(request, f'Добавлено {amount} ₽ к цели «{goal.name}»')

            if goal.target_amount and goal.target_amount > 0:
                goal.progress_percentage = min(100.0, (float(goal.current_amount) / float(goal.target_amount)) * 100)
            else:
                goal.progress_percentage = 0
            goal.save()

            return redirect(redirect_to)

        except (ValueError, TypeError):
            messages.error(request, 'Введите корректную сумму целым числом (например: 1000)')
            return redirect(redirect_to)
        except Exception as e:
            messages.error(request, f'Произошла ошибка: {str(e)}')
            return redirect(redirect_to)

    return redirect(redirect_to)


@login_required
def family_list(request):
    """Список семей пользователя"""
    families = Family.objects.filter(created_by=request.user) | Family.objects.filter(members__user=request.user)
    families = families.distinct()
    families_with_avatars = [(f, _family_avatar_url(f)) for f in families]
    return render(request, 'finance/family_list.html', {'families_with_avatars': families_with_avatars})


@login_required
def notifications_list(request):
    """Список уведомлений (непрочитанные отображаются первыми и выделены)."""
    # Непрочитанные первыми (is_read False < True, поэтому order_by('is_read') даёт unread первыми)
    notifications = Notification.objects.filter(user=request.user).order_by('is_read', '-created_at')[:50]
    return render(request, 'finance/notifications_list.html', {'notifications': notifications})


@login_required
def notifications_mark_all_read(request):
    """Отметить все уведомления как прочитанные."""
    if request.method == 'POST':
        from django.utils import timezone
        updated = Notification.objects.filter(user=request.user, is_read=False).update(is_read=True, read_at=timezone.now())
        messages.success(request, f'Отмечено прочитанными: {updated}')
    return redirect('notifications_list')


@login_required
def profile_edit(request):
    """Редактирование профиля"""
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Профиль обновлён!')
            return redirect('profile_edit')
    else:
        form = ProfileUpdateForm(instance=request.user)
    # Безопасный URL аватара — только если файл существует (избегаем 404)
    avatar_url = None
    if request.user.avatar:
        try:
            from django.core.files.storage import default_storage
            if default_storage.exists(request.user.avatar.name):
                avatar_url = request.user.avatar.url
        except Exception:
            pass
    return render(request, 'finance/profile_edit.html', {'form': form, 'avatar_url': avatar_url})


@login_required
def family_create(request):
    """Создание семьи"""
    from .forms import FamilyForm
    if request.method == 'POST':
        form = FamilyForm(request.POST)
        if form.is_valid():
            family = form.save(commit=False)
            family.created_by = request.user
            family.family_currency = 'RUB'
            family.save()
            FamilyMember.objects.create(family=family, user=request.user, role='creator')
            messages.success(request, 'Семья создана!')
            return redirect('family_list')
    else:
        form = FamilyForm()
    families = Family.objects.filter(created_by=request.user) | Family.objects.filter(members__user=request.user).distinct()
    families_with_avatars = [(f, _family_avatar_url(f)) for f in families]
    return render(request, 'finance/family_list.html', {'form': form, 'families_with_avatars': families_with_avatars, 'show_create_form': True})


def _family_avatar_url(family):
    """Безопасный URL аватара семьи (только если файл есть)."""
    if not family.avatar:
        return None
    try:
        from django.core.files.storage import default_storage
        if default_storage.exists(family.avatar.name):
            return family.avatar.url
    except Exception:
        pass
    return None


@login_required
def family_detail(request, family_id):
    """Детали семьи"""
    family = get_object_or_404(Family, id=family_id)
    if family.created_by != request.user and not family.members.filter(user=request.user).exists():
        messages.error(request, 'Нет доступа к этой семье')
        return redirect('family_list')
    members = family.members.select_related('user').all()
    is_creator = family.created_by == request.user
    is_admin = is_creator or family.members.filter(user=request.user, role='admin').exists()
    can_invite = _user_can_invite_to_family(request, family)
    can_create_goal = _user_can_create_family_goals(request, family)
    family_goals = FinancialGoal.objects.filter(family=family).order_by('deadline')
    family_avatar_url = _family_avatar_url(family)
    members_with_avatars = []
    try:
        from django.core.files.storage import default_storage
        for m in members:
            url = None
            if m.user.avatar and default_storage.exists(m.user.avatar.name):
                url = m.user.avatar.url
            members_with_avatars.append((m, url))
    except Exception:
        members_with_avatars = [(m, None) for m in members]
    if not members_with_avatars:
        members_with_avatars = [(m, None) for m in members]
    # График пополнений по месяцам — отдельно по каждой цели (датасеты по целям)
    from django.utils import formats
    from django.db.models import Min, Max
    from collections import defaultdict
    chart_month_from = request.GET.get('month_from', '').strip()
    chart_month_to = request.GET.get('month_to', '').strip()
    contributions_qs = GoalContribution.objects.filter(goal__family=family)
    if chart_month_from:
        try:
            y, m = int(chart_month_from[:4]), int(chart_month_from[5:7])
            start_date = date(y, m, 1)
            contributions_qs = contributions_qs.filter(contributed_at__date__gte=start_date)
        except (ValueError, IndexError):
            pass
    if chart_month_to:
        try:
            y, m = int(chart_month_to[:4]), int(chart_month_to[5:7])
            _, last_day = calendar.monthrange(y, m)
            end_date = date(y, m, last_day)
            contributions_qs = contributions_qs.filter(contributed_at__date__lte=end_date)
        except (ValueError, IndexError):
            pass
    contributions_raw = contributions_qs.annotate(
        month=TruncMonth('contributed_at')
    ).values('goal_id', 'goal__name', 'month').annotate(
        total=Sum('amount')
    ).order_by('month')
    # Список месяцев для выбора в фильтре
    agg = GoalContribution.objects.filter(goal__family=family).aggregate(
        first=Min('contributed_at'), last=Max('contributed_at'))
    chart_available_months = []
    if agg['first'] and agg['last']:
        cur = agg['first'].date().replace(day=1)
        end = agg['last'].date().replace(day=1)
        while cur <= end:
            chart_available_months.append((cur.strftime('%Y-%m'), formats.date_format(cur, 'F Y')))
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)
    else:
        today = timezone.now().date()
        for i in range(11, -1, -1):
            y, m = today.year, today.month - i
            while m <= 0:
                m += 12
                y -= 1
            d = date(y, m, 1)
            chart_available_months.append((d.strftime('%Y-%m'), formats.date_format(d, 'F Y')))
    # Собираем все месяцы (последние 12) и по каждой цели — сумму за месяц
    today = timezone.now().date()
    chart_labels = []
    for i in range(11, -1, -1):
        y, m = today.year, today.month - i
        while m <= 0:
            m += 12
            y -= 1
        d = date(y, m, 1)
        chart_labels.append(d.strftime('%Y-%m'))
    by_goal = defaultdict(dict)  # goal_id -> { month_str: total }
    for row in contributions_raw:
        month_val = row.get('month')
        if month_val:
            month_str = formats.date_format(month_val, 'Y-m')
            gid = row.get('goal_id')
            gname = row.get('goal__name') or 'Цель'
            if gid not in by_goal:
                by_goal[gid] = {'name': gname, 'months': {}}
            by_goal[gid]['months'][month_str] = float(row.get('total') or 0)
    chart_datasets = []
    colors = [
        'rgba(67, 97, 238, 0.8)', 'rgba(34, 197, 94, 0.8)', 'rgba(234, 88, 12, 0.8)',
        'rgba(168, 85, 247, 0.8)', 'rgba(14, 165, 233, 0.8)', 'rgba(225, 29, 72, 0.8)',
    ]
    for i, (gid, info) in enumerate(by_goal.items()):
        data = [info['months'].get(m, 0) for m in chart_labels]
        chart_datasets.append({
            'label': info['name'][:30],
            'data': data,
            'backgroundColor': colors[i % len(colors)],
            'borderColor': colors[i % len(colors)].replace('0.8', '1'),
            'borderWidth': 1,
        })
    chart_labels_json = json.dumps(chart_labels)
    chart_datasets_json = json.dumps(chart_datasets)
    # История пополнений: кто, когда, сколько, по какой цели
    contributions_history = GoalContribution.objects.filter(
        goal__family=family
    ).select_related('user', 'goal').order_by('-contributed_at')[:100]
    # Сумма пополнения каждого пользователя по каждой семейной цели
    from django.db.models import Sum as DbSum
    contributions_by_goal_user = GoalContribution.objects.filter(
        goal__family=family
    ).values('goal_id', 'user_id').annotate(total=DbSum('amount')).order_by('goal_id', 'user_id')
    goal_contributions_by_user = {}  # goal_id -> [(display_name, total), ...]
    for row in contributions_by_goal_user:
        gid = row['goal_id']
        if gid not in goal_contributions_by_user:
            goal_contributions_by_user[gid] = []
        try:
            u = CustomUser.objects.get(id=row['user_id']) if row['user_id'] else None
            display = (u.get_full_name() or u.username) if u else '—'
        except CustomUser.DoesNotExist:
            display = '—'
        goal_contributions_by_user[gid].append((display, float(row['total'] or 0)))
    family_goals_with_contributions = [
        (goal, goal_contributions_by_user.get(goal.id, []))
        for goal in family_goals
    ]
    return render(request, 'finance/family_detail.html', {
        'chart_labels_json': chart_labels_json,
        'chart_datasets_json': chart_datasets_json,
        'chart_month_from': chart_month_from,
        'chart_month_to': chart_month_to,
        'chart_available_months': chart_available_months,
        'family': family, 'members_with_avatars': members_with_avatars, 'is_creator': is_creator, 'is_admin': is_admin,
        'can_invite': can_invite, 'can_create_goal': can_create_goal,
        'family_goals': family_goals, 'family_goals_with_contributions': family_goals_with_contributions,
        'pending_emails': [], 'family_avatar_url': family_avatar_url,
        'chart_labels': chart_labels,
        'contributions_history': contributions_history,
    })


def _user_can_invite_to_family(request, family):
    """Может ли пользователь приглашать в семью: создатель, админ или участник с правом members_can_invite."""
    if family.created_by == request.user:
        return True
    if family.members.filter(user=request.user, role='admin').exists():
        return True
    if family.members.filter(user=request.user).exists() and getattr(family, 'members_can_invite', False):
        return True
    return False


def _user_can_create_family_goals(request, family):
    """Может ли пользователь создавать семейные цели."""
    if family.created_by == request.user:
        return True
    if family.members.filter(user=request.user, role='admin').exists():
        return True
    if getattr(family, 'members_can_create_goals', True):
        return True
    return False


@login_required
def family_get_invite_link(request, family_id):
    """Возвращает ссылку приглашения в семью (JSON для AJAX)."""
    family = get_object_or_404(Family, id=family_id)
    if family.created_by != request.user and not family.members.filter(user=request.user).exists():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
        return redirect('family_detail', family_id=family_id)
    if not _user_can_invite_to_family(request, family):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Только создатель/админ или участники с правом приглашения могут копировать ссылку. Включите в настройках семьи.'}, status=403)
        messages.error(request, 'Только создатель/админ или участники с правом приглашения могут копировать ссылку.')
        return redirect('family_detail', family_id=family_id)
    from django.core.signing import Signer
    from urllib.parse import quote
    signer = Signer()
    token = quote(signer.sign(str(family_id)), safe='')
    link = request.build_absolute_uri(reverse('family_accept_invite', kwargs={'token': token}))
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'link': link})
    messages.success(request, 'Ссылка скопирована в буфер обмена')
    return redirect('family_detail', family_id=family_id)


@login_required
def family_invite(request, family_id):
    """Приглашение по email: создаёт FamilyInvitation и уведомление приглашённому."""
    family = get_object_or_404(Family, id=family_id)
    if not _user_can_invite_to_family(request, family):
        messages.error(request, 'У вас нет права приглашать в семью. Включите в настройках семьи.')
        return redirect('family_detail', family_id=family_id)
    if request.method == 'POST':
        email = (request.POST.get('email') or '').strip().lower()
        if not email:
            messages.error(request, 'Введите email для приглашения.')
            return redirect('family_detail', family_id=family_id)
        from django.core.signing import Signer
        from urllib.parse import quote
        signer = Signer()
        token = quote(signer.sign(str(family_id)), safe='')
        invite_link = request.build_absolute_uri(reverse('family_accept_invite', kwargs={'token': token}))
        # Создаём приглашение (для истории)
        inv, _ = FamilyInvitation.objects.get_or_create(
            family=family, invitee_email=email, status='pending',
            defaults={'inviter': request.user, 'token': token[:100], 'expires_at': timezone.now() + timedelta(days=7), 'proposed_role': 'member', 'proposed_permission': 'full'}
        )
        if not _:
            inv.token = token[:100]
            inv.expires_at = timezone.now() + timedelta(days=7)
            inv.save(update_fields=['token', 'expires_at'])
        # Уведомление приглашённому — если пользователь с таким email есть
        invitee_user = None
        try:
            from .models import CustomUser
            invitee_user = CustomUser.objects.get(email__iexact=email)
        except CustomUser.DoesNotExist:
            pass
        if invitee_user:
            Notification.objects.create(
                user=invitee_user,
                notification_type='family_invite',
                title='Приглашение в семью',
                message=f'Вас пригласили в семью «{family.name}». Перейдите по ссылке, чтобы присоединиться.',
                data={'family_id': str(family_id), 'invite_token': token, 'inviter': request.user.username},
            )
            messages.success(request, f'Приглашение отправлено на {email}. Пользователь увидит уведомление.')
        else:
            messages.success(request, f'Ссылка для приглашения создана. Отправьте её вручную на {email}: {invite_link[:80]}...')
    return redirect('family_detail', family_id=family_id)


@login_required
def family_remove_member(request, family_id):
    """Удаление участника из семьи (только создатель)."""
    family = get_object_or_404(Family, id=family_id)
    if family.created_by != request.user:
        messages.error(request, 'Только создатель семьи может удалять участников.')
        return redirect('family_detail', family_id=family_id)
    if request.method == 'POST':
        member_id = request.POST.get('member_id')
        if member_id:
            member = family.members.filter(id=member_id).first()
            if member and member.role != 'creator':
                member.delete()
                messages.success(request, 'Участник удалён из семьи.')
            elif member and member.role == 'creator':
                messages.error(request, 'Нельзя удалить создателя семьи.')
        else:
            messages.error(request, 'Укажите участника.')
    return redirect('family_detail', family_id=family_id)


def family_accept_invite(request, token):
    """Принятие приглашения в семью по ссылке."""
    from urllib.parse import unquote
    from django.core.signing import Signer, BadSignature
    if not request.user.is_authenticated:
        messages.info(request, 'Войдите, чтобы присоединиться к семье')
        return redirect('auth')
    try:
        raw = unquote(token)
        family_id = Signer().unsign(raw)
        family = get_object_or_404(Family, id=family_id)
    except (BadSignature, ValueError, Family.DoesNotExist):
        messages.error(request, 'Ссылка приглашения недействительна или устарела')
        return redirect('family_list')
    if family.members.filter(user=request.user).exists():
        messages.info(request, 'Вы уже в этой семье')
        return redirect('family_detail', family_id=family_id)
    FamilyMember.objects.get_or_create(family=family, user=request.user, defaults={'role': 'member'})
    # Уведомление создателю семьи: кто присоединился
    display_name = request.user.get_full_name() or request.user.username
    Notification.objects.create(
        user=family.created_by,
        notification_type='member_joined',
        title='Новый участник в семье',
        message=f'{display_name} присоединился(ась) к семье «{family.name}».',
        data={'family_id': str(family_id), 'user_id': str(request.user.id), 'username': request.user.username},
    )
    messages.success(request, f'Вы присоединились к семье «{family.name}»')
    return redirect('family_detail', family_id=family_id)


@login_required
def family_admin_chart(request, family_id):
    """График пополнений по целям по месяцам — доступен всем участникам семьи."""
    family = get_object_or_404(Family, id=family_id)
    if family.created_by != request.user and not family.members.filter(user=request.user).exists():
        messages.error(request, 'Нет доступа к этой семье')
        return redirect('family_list')
    from django.utils import formats
    from django.db.models import Min, Max
    from collections import defaultdict
    chart_month_from = request.GET.get('month_from', '').strip()
    chart_month_to = request.GET.get('month_to', '').strip()
    contributions_qs = GoalContribution.objects.filter(goal__family=family)
    if chart_month_from:
        try:
            y, m = int(chart_month_from[:4]), int(chart_month_from[5:7])
            start_date = date(y, m, 1)
            contributions_qs = contributions_qs.filter(contributed_at__date__gte=start_date)
        except (ValueError, IndexError):
            pass
    if chart_month_to:
        try:
            y, m = int(chart_month_to[:4]), int(chart_month_to[5:7])
            _, last_day = calendar.monthrange(y, m)
            end_date = date(y, m, last_day)
            contributions_qs = contributions_qs.filter(contributed_at__date__lte=end_date)
        except (ValueError, IndexError):
            pass
    contributions_raw = contributions_qs.annotate(
        month=TruncMonth('contributed_at')
    ).values('goal_id', 'goal__name', 'month').annotate(
        total=Sum('amount')
    ).order_by('month')
    months_set = set()
    by_goal = defaultdict(lambda: {'name': '', 'months': {}})
    for row in contributions_raw:
        month_val = row.get('month')
        if month_val:
            month_str = formats.date_format(month_val, 'Y-m')
            months_set.add(month_str)
            gid = row.get('goal_id')
            by_goal[gid]['name'] = row.get('goal__name') or 'Цель'
            by_goal[gid]['months'][month_str] = float(row.get('total') or 0)
    chart_labels = sorted(months_set) if months_set else []
    chart_datasets = []
    colors = [
        'rgba(67, 97, 238, 0.8)', 'rgba(34, 197, 94, 0.8)', 'rgba(234, 88, 12, 0.8)',
        'rgba(168, 85, 247, 0.8)', 'rgba(14, 165, 233, 0.8)', 'rgba(225, 29, 72, 0.8)',
    ]
    for i, (gid, info) in enumerate(by_goal.items()):
        data = [info['months'].get(m, 0) for m in chart_labels]
        chart_datasets.append({
            'label': (info['name'] or 'Цель')[:30],
            'data': data,
            'backgroundColor': colors[i % len(colors)],
            'borderColor': colors[i % len(colors)].replace('0.8', '1'),
            'borderWidth': 1,
        })
    agg = GoalContribution.objects.filter(goal__family=family).aggregate(
        first=Min('contributed_at'), last=Max('contributed_at'))
    chart_available_months = []
    if agg['first'] and agg['last']:
        cur = agg['first'].date().replace(day=1)
        end = agg['last'].date().replace(day=1)
        while cur <= end:
            chart_available_months.append((cur.strftime('%Y-%m'), formats.date_format(cur, 'F Y')))
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)
    else:
        today = timezone.now().date()
        for i in range(11, -1, -1):
            y, m = today.year, today.month - i
            while m <= 0:
                m += 12
                y -= 1
            d = date(y, m, 1)
            chart_available_months.append((d.strftime('%Y-%m'), formats.date_format(d, 'F Y')))
    return render(request, 'finance/family_admin_chart.html', {
        'family': family,
        'chart_labels_json': json.dumps(chart_labels),
        'chart_datasets_json': json.dumps(chart_datasets),
        'chart_month_from': chart_month_from,
        'chart_month_to': chart_month_to,
        'chart_available_months': chart_available_months,
        'family_avatar_url': _family_avatar_url(family),
    })


@login_required
def family_settings(request, family_id):
    """Настройки семьи (только создатель/админ)."""
    family = get_object_or_404(Family, id=family_id)
    is_admin = family.created_by == request.user or family.members.filter(user=request.user, role='admin').exists()
    if not is_admin:
        messages.error(request, 'Нет доступа к настройкам')
        return redirect('family_detail', family_id=family_id)
    if request.method == 'POST':
        family.name = request.POST.get('name', family.name)[:100]
        family.members_can_create_goals = request.POST.get('members_can_create_goals') == 'on'
        family.members_can_invite = request.POST.get('members_can_invite') == 'on'
        if request.FILES.get('avatar'):
            family.avatar = request.FILES['avatar']
        family.save()
        messages.success(request, 'Настройки сохранены')
        return redirect('family_detail', family_id=family_id)
    return render(request, 'finance/family_settings.html', {
        'family': family, 'family_avatar_url': _family_avatar_url(family),
    })


@login_required
def family_member_display_name(request, family_id, member_id):
    """Изменение отображаемого имени в семье (для себя или участника — создатель/админ)."""
    family = get_object_or_404(Family, id=family_id)
    member = get_object_or_404(FamilyMember, id=member_id, family=family)
    can_edit = (member.user == request.user or
                family.created_by == request.user or
                family.members.filter(user=request.user, role='admin').exists())
    if not can_edit:
        messages.error(request, 'Нет доступа')
        return redirect('family_detail', family_id=family_id)
    if request.method == 'POST':
        member.display_name = (request.POST.get('display_name') or '')[:80]
        member.save()
        messages.success(request, 'Имя обновлено')
    return redirect('family_detail', family_id=family_id)


@login_required
def family_goal_create(request, family_id):
    """Создание семейной цели — редирект на дашборд с формой цели."""
    family = get_object_or_404(Family, id=family_id)
    if family.created_by != request.user and not family.members.filter(user=request.user).exists():
        messages.error(request, 'Нет доступа')
        return redirect('family_list')
    if not _user_can_create_family_goals(request, family):
        messages.error(request, 'В этой семье только создатель/админ могут создавать цели. Включите опцию в настройках семьи.')
        return redirect('family_detail', family_id=family_id)
    return redirect(reverse('dashboard') + '?tab=goals&family_id=' + str(family_id))