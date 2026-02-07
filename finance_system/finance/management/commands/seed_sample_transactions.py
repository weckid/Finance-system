"""
Создаёт 50–60 тестовых транзакций за январь–февраль для нескольких пользователей.
Формат как в шаблоне Excel: Дата | Сумма | Категория | Магазин.
"""
import os
import random
from datetime import datetime, timedelta

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils import timezone

from finance.models import Account, Category, CustomUser, Transaction


SAMPLE_DATA = [
    (150, 2500, 'Еда', 'Пятёрочка'),
    (200, 890, 'Еда', 'Магнит'),
    (150, 450, 'Кафе и рестораны', 'Шоколадница'),
    (350, 1200, 'Кафе и рестораны', 'Макдоналдс'),
    (100, 320, 'Еда', 'Пятёрочка'),
    (500, 1800, 'Здоровье', 'Аптека 36.6'),
    (250, 650, 'Транспорт', 'Заправка'),
    (150, 550, 'Еда', 'Перекрёсток'),
    (80, 280, 'Еда', 'Пятёрочка'),
    (400, 1400, 'Развлечения', 'Кинотеатр'),
    (200, 120, 'Связь', 'МТС'),
    (300, 2400, 'Одежда и обувь', 'H&M'),
    (180, 720, 'Еда', 'ВкусВилл'),
    (220, 450, 'Кафе и рестораны', 'Starbucks'),
    (90, 190, 'Еда', 'Пятёрочка'),
    (350, 1100, 'Транспорт', 'Яндекс.Такси'),
    (150, 380, 'Еда', 'Дикси'),
    (280, 900, 'Здоровье', 'Столичка'),
    (120, 520, 'Еда', 'Магнит'),
    (450, 1650, 'Развлечения', 'Ресторан'),
    (100, 250, 'Еда', 'Пятёрочка'),
    (200, 780, 'Коммунальные услуги', 'ЖКХ'),
    (170, 420, 'Еда', 'Лента'),
    (190, 590, 'Кафе и рестораны', 'Кофе Хауз'),
    (130, 310, 'Еда', 'Пятёрочка'),
    (500, 2200, 'Одежда и обувь', 'Zara'),
    (160, 680, 'Транспорт', 'Метро'),
    (140, 390, 'Еда', 'Перекрёсток'),
    (210, 850, 'Здоровье', 'Ригла'),
    (110, 270, 'Еда', 'Магнит'),
    (380, 1350, 'Развлечения', 'Концерт'),
    (95, 220, 'Еда', 'Пятёрочка'),
    (260, 950, 'Товары для дома', 'Икеа'),
    (175, 410, 'Еда', 'ВкусВилл'),
    (230, 620, 'Кафе и рестораны', 'Додо Пицца'),
    (105, 290, 'Еда', 'Пятёрочка'),
    (420, 1800, 'Образование', 'Онлайн-курсы'),
    (155, 480, 'Еда', 'Магнит'),
    (195, 540, 'Транспорт', 'Сбербанк топливо'),
    (125, 330, 'Еда', 'Пятёрочка'),
    (340, 1250, 'Развлечения', 'Театр'),
    (165, 460, 'Еда', 'Дикси'),
    (205, 710, 'Здоровье', 'Аптека'),
    (115, 260, 'Еда', 'Пятёрочка'),
    (270, 880, 'Кафе и рестораны', 'Ресторан'),
    (145, 370, 'Еда', 'Перекрёсток'),
    (310, 1100, 'Одежда и обувь', 'Ozon'),
    (135, 350, 'Еда', 'Магнит'),
    (185, 510, 'Транспорт', 'Яндекс'),
    (118, 285, 'Еда', 'Пятёрочка'),
    (390, 1420, 'Развлечения', 'Боулинг'),
    (152, 430, 'Еда', 'ВкусВилл'),
    (215, 660, 'Кафе и рестораны', 'KFC'),
    (128, 305, 'Еда', 'Пятёрочка'),
    (240, 820, 'Коммунальные услуги', 'Электричество'),
    (162, 445, 'Еда', 'Лента'),
    (198, 570, 'Здоровье', 'Клиника'),
    (122, 295, 'Еда', 'Пятёрочка'),
]


class Command(BaseCommand):
    help = 'Создаёт 50–60 тестовых транзакций (январь–февраль) для нескольких пользователей.'

    def add_arguments(self, parser):
        parser.add_argument('--excel', type=str, help='Путь для сохранения Excel файла.')
        parser.add_argument('--no-db', action='store_true', help='Только создать Excel, не писать в БД.')

    def handle(self, *args, **options):
        excel_path = options.get('excel')
        no_db = options.get('no_db', False)

        users = list(CustomUser.objects.all()[:10])
        user_labels = [u.username for u in users] if users else ['user1', 'user2', 'user3', 'admin']

        categories = {c.name: c for c in Category.objects.filter(type='expense')}
        start = datetime(2025, 1, 1)
        end = datetime(2025, 2, 28)

        rows = []
        for i, (min_amt, max_amt, cat_name, merchant) in enumerate(SAMPLE_DATA[:58]):
            day_offset = random.randint(0, (end - start).days)
            dt = start + timedelta(days=day_offset)
            amount = round(random.uniform(min_amt, max_amt), 2)
            label = random.choice(user_labels)
            user = users[user_labels.index(label)] if users and label in user_labels else None
            rows.append((dt, amount, cat_name, merchant, label, user))

        rows.sort(key=lambda r: r[0])

        if not no_db and users:
            for cat_name in {r[2] for r in rows if r[2]}:
                if cat_name not in categories:
                    c, _ = Category.objects.get_or_create(
                        name=cat_name, type='expense', is_system=False, owner=None,
                        defaults={'color': '#4ECDC4'}
                    )
                    categories[cat_name] = c

            created = 0
            for dt, amount, cat_name, merchant, _label, user in rows:
                if not user:
                    continue
                account = Account.objects.filter(owner=user, is_active=True).first()
                if not account:
                    account = Account.objects.create(
                        owner=user, name='Основной счёт', account_type='debit',
                        ownership='personal', currency='RUB', is_active=True
                    )
                category = categories.get(cat_name) if cat_name else None
                Transaction.objects.create(
                    user=user, account=account, category=category,
                    amount=amount, type='expense', currency='RUB',
                    merchant=merchant or None,
                    date=timezone.make_aware(dt.replace(hour=random.randint(8, 20), minute=random.randint(0, 59))),
                    created_via='import'
                )
                created += 1
            self.stdout.write(self.style.SUCCESS(f'Создано транзакций: {created}'))
        elif not no_db and not users:
            self.stdout.write(self.style.WARNING('Нет пользователей. Только Excel будет создан.'))

        out_path = excel_path or os.path.join(settings.BASE_DIR, 'transactions_sample_50.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Транзакции'
        headers = ['Дата', 'Сумма', 'Категория', 'Магазин/Продавец', 'Пользователь']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, (dt, amount, cat_name, merchant, label, _) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=dt.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=amount)
            ws.cell(row=row_idx, column=3, value=cat_name)
            ws.cell(row=row_idx, column=4, value=merchant)
            ws.cell(row=row_idx, column=5, value=label)
        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(f'Excel сохранён: {out_path}'))
